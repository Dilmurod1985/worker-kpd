from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file
import os
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
import io
from datetime import datetime
from collections import defaultdict
from database import init_db, add_record, get_all_records, get_records_by_date

app = Flask(__name__)

# Инициализация базы данных
init_db()

# Папка для временной загрузки фото
app.config['UPLOAD_FOLDER'] = 'uploads'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Список товаров (твои реальные позиции)
PRODUCTS = [
    "DONER 100 QYMA",
    "DONER 50/70",
    "DONER 30/70",
    "DONER 50/50 BELORUS",
    "DONER 50/50 CITY",
    "DONER 50/50 SEVIMLI",
    "DONER 50/50 YUMA",
    "DONER 60/40 YUMA",
    "DONER 70/30",
    "DONER 70/30 YUMA",
    "DONER TOVUQ",
    "OQTEPA LAVASH 50/50",
    "DONER WENDY'S (60/50) DONA",
    "DONER BURGER ХОЖИАҚБАР (10) 50/50",
    "DONER BURGER ХОЖИАҚБАР (20) 50/50 D",
    "DONER BURGER ХОЖИАҚБАР (30) 50/50",
    "DONER BURGER ХОЖИАҚБАР (40) 50/50 D",
    "DONER BURGER ХОЖИАҚБАР (5) 50/50",
    "DONER MISSIONFOODS (20) 100 QYMA",
    "DONER MISSIONFOODS (50) 100 QYMA D",
    "DONER MISSIONFOODS (60) 100 DONA",
    "DONER MISSIONFOODS (60) 50/50 DONA",
    "Сихлаш",
    "Фарш говяжий"
]

# Группы товаров и их ассортимент
PRODUCT_GROUPS = {
    "ALLMAKON GROUP": [
        "DONER 100 QYMA",
        "DONER 50/70",
        "DONER 30/70",
        "DONER 50/50 BELORUS",
        "DONER 50/50 CITY",
        "DONER 50/50 SEVIMLI",
        "DONER 50/50 YUMA",
        "DONER 60/40 YUMA",
        "DONER 70/30 YUMA"
    ],
    "OQTEPA LAVASH": [
        "OQTEPA LAVASH 50/50",
        "DONER 50/70",
        "DONER 70/30",
        "DONER TOVUQ"
    ],
    "WENDY'S": [
        "DONER WENDY'S (60/50) DONA"
    ],
    "BURGER XOJIAKBAR": [
        "DONER BURGER ХОЖИАҚБАР (10) 50/50",
        "DONER BURGER ХОЖИАҚБАР (20) 50/50 D",
        "DONER BURGER ХОЖИАҚБАР (30) 50/50",
        "DONER BURGER ХОЖИАҚБАР (40) 50/50 D",
        "DONER BURGER ХОЖИАҚБАР (5) 50/50"
    ],
    "MISSION FOODS": [
        "DONER MISSIONFOODS (20) 100 QYMA",
        "DONER MISSIONFOODS (50) 100 QYMA D",
        "DONER MISSIONFOODS (60) 100 DONA",
        "DONER MISSIONFOODS (60) 50/50 DONA"
    ]
}

GROUPS_LIST = list(PRODUCT_GROUPS.keys())

# Актуальный список работников с отделами
WORKERS_TABLE = [
    # Сихлаш (обновлено)
    {"id": "57", "fio": "Alimov Murod", "otdel": "Сихлаш", "category": "4 категория", "oklad": 4200000},
    {"id": "85", "fio": "Aliqulov Diyor", "otdel": "Сихлаш", "category": "5 категория", "oklad": 3640000},
    {"id": "3", "fio": "Balqiboev Asadbek", "otdel": "Сихлаш", "category": "4 категория", "oklad": 4200000},
    {"id": "5", "fio": "Balqiboev Temur", "otdel": "Сихлаш", "category": "2 категория", "oklad": 6160000},
    {"id": "76", "fio": "Boboqulov Ma'rufjon", "otdel": "Сихлаш", "category": "4 категория", "oklad": 4200000},
    {"id": "77", "fio": "Botiraliyev Iskandar", "otdel": "Сихлаш", "category": "4 категория", "oklad": 4200000},
    {"id": "86", "fio": "Botirov Ulug'bek", "otdel": "Сихлаш", "category": "4 категория", "oklad": 4200000},
    {"id": "74", "fio": "Hakimov Islom", "otdel": "Сихлаш", "category": "2 категория", "oklad": 6160000},
    {"id": "41", "fio": "Ibragimov Ozodbek", "otdel": "Сихлаш", "category": "3 категория", "oklad": 4900000},
    {"id": "67", "fio": "Jonizaqov Xasanjon", "otdel": "Сихлаш", "category": "4 категория", "oklad": 4200000},
    {"id": "101", "fio": "Jonizaqov Xusanjon", "otdel": "Сихлаш", "category": "4 категория", "oklad": 4200000},
    {"id": "64", "fio": "Keldibekov Muxammadkodir", "otdel": "Сихлаш", "category": "2 категория", "oklad": 6160000},
    {"id": "46", "fio": "Kuchkinov Abdulaziz", "otdel": "Сихлаш", "category": "2 категория", "oklad": 6160000},
    {"id": "84", "fio": "Lapasov Javohir", "otdel": "Сихлаш", "category": "4 категория", "oklad": 4200000},
    {"id": "102", "fio": "Maxkamov Jahongir", "otdel": "Сихлаш", "category": "стаж", "oklad": 3500000},
    {"id": "28", "fio": "Miraliev Mirxasan", "otdel": "Сихлаш", "category": "2 категория", "oklad": 6160000},
    {"id": "14", "fio": "Miraliev Mirxusan", "otdel": "Сихлаш", "category": "4 категория", "oklad": 4200000},
    {"id": "57", "fio": "Muxammadiev Abbos", "otdel": "Сихлаш", "category": "2 категория", "oklad": 6160000},
    {"id": "52", "fio": "Nazarov Abdukodir", "otdel": "Сихлаш", "category": "2 категория", "oklad": 6160000},
    {"id": "49", "fio": "Nishonov Maxmudjon", "otdel": "Сихлаш", "category": "1 категория", "oklad": 6440000},
    {"id": "6", "fio": "Normatov Avazbek", "otdel": "Сихлаш", "category": "3 категория", "oklad": 4900000},
    {"id": "36", "fio": "O'ktamov Sardor", "otdel": "Сихлаш", "category": "4 категория", "oklad": 4200000},
    {"id": "103", "fio": "Olimxonov Nurmuhammad", "otdel": "Сихлаш", "category": "5 категория", "oklad": 3640000},
    {"id": "24", "fio": "Ortiqboev Ma'rufjon", "otdel": "Сихлаш", "category": "4 категория", "oklad": 4200000},
    {"id": "2", "fio": "Qo'chqorov Sardor", "otdel": "Сихлаш", "category": "2 категория", "oklad": 6160000},
    {"id": "10", "fio": "Qo'shbokov Inomjon", "otdel": "Сихлаш", "category": "3 категория", "oklad": 4900000},
    {"id": "30", "fio": "Raimberdiyev Dilshod", "otdel": "Сихлаш", "category": "5 категория", "oklad": 3640000},
    {"id": "104", "fio": "Rustamjonov Aziz", "otdel": "Сихлаш", "category": "4 категория", "oklad": 4200000},
    {"id": "58", "fio": "Sattorov Ulug'bek", "otdel": "Сихлаш", "category": "5 категория", "oklad": 3640000},
    {"id": "35", "fio": "Shermatov Javlon", "otdel": "Сихлаш", "category": "2 категория", "oklad": 6160000},
    {"id": "73", "fio": "Tashmatov Raximjon", "otdel": "Сихлаш", "category": "5 категория", "oklad": 3640000},
    {"id": "26", "fio": "To'xtamurodov Izzatilla", "otdel": "Сихлаш", "category": "4 категория", "oklad": 4200000},
    {"id": "51", "fio": "Tugalov Sherzod", "otdel": "Сихлаш", "category": "3 категория", "oklad": 4900000},
    {"id": "105", "fio": "Vaxobjonov Avazbek", "otdel": "Сихлаш", "category": "5 категория", "oklad": 3640000},
    {"id": "1", "fio": "Xalilov Nodirbek", "otdel": "Сихлаш", "category": "1 категория", "oklad": 6440000},
    {"id": "27", "fio": "Xapizov Ozodbek", "otdel": "Сихлаш", "category": "5 категория", "oklad": 3640000},
    {"id": "39", "fio": "Xayrullaev Jasurbek", "otdel": "Сихлаш", "category": "3 категория", "oklad": 4900000},
    {"id": "4", "fio": "Xoliqov Jumanazarbek", "otdel": "Сихлаш", "category": "3 категория", "oklad": 4900000},
    {"id": "61", "fio": "Obidov Ziyodulla", "otdel": "Сихлаш", "category": "3 категория", "oklad": 4900000},
    {"id": "75", "fio": "Turg'unboyev Asadbek", "otdel": "Сихлаш", "category": "стаж", "oklad": 3500000},

    # Разделка
    {"id": "42", "fio": "Abdullayev Sherzod", "otdel": "Разделка", "category": "5 категория", "oklad": 3640000},
    {"id": "43", "fio": "Abduvaliyev Yaxyo", "otdel": "Разделка", "category": "5 категория", "oklad": 3640000},
    {"id": "44", "fio": "Bo'riqulov Shaxzod", "otdel": "Разделка", "category": "2 категория", "oklad": 6160000},
    {"id": "45", "fio": "Keldiev Azizbek", "otdel": "Разделка", "category": "4 категория", "oklad": 4200000},
    {"id": "46", "fio": "Xurramov Shamsiddin", "otdel": "Разделка", "category": "4 категория", "oklad": 4200000},
    {"id": "47", "fio": "Abdulxamidov Xasan", "otdel": "Разделка", "category": "1 категория", "oklad": 6440000},
    {"id": "48", "fio": "Eshmurodov Mirzoxid", "otdel": "Разделка", "category": "2 категория", "oklad": 6160000},
    {"id": "49", "fio": "Kamilov Rustam", "otdel": "Разделка", "category": "4 категория", "oklad": 4200000},
    {"id": "50", "fio": "Mirzakulov Mirolim", "otdel": "Разделка", "category": "1 категория", "oklad": 6440000},
    {"id": "51", "fio": "Muxtorov Saidali", "otdel": "Разделка", "category": "5 категория", "oklad": 3640000},
    {"id": "52", "fio": "Mirzaaxmedov Qochqor", "otdel": "Разделка", "category": "4 категория", "oklad": 4200000},
    {"id": "53", "fio": "Kanaev Murodjon", "otdel": "Разделка", "category": "5 категория", "oklad": 3640000},
    {"id": "54", "fio": "Saidxonov Nasriddin", "otdel": "Разделка", "category": "5 категория", "oklad": 3640000},
    {"id": "55", "fio": "Nosirov Adhamjon", "otdel": "Разделка", "category": "5 категория", "oklad": 3640000},
    {"id": "56", "fio": "Urolov Xumoyun", "otdel": "Разделка", "category": "5 категория", "oklad": 3640000},
    {"id": "57", "fio": "Jo'raboyev Boymurod", "otdel": "Разделка", "category": "стаж", "oklad": 3500000},
    {"id": "58", "fio": "G'aniyev Muzaffar", "otdel": "Разделка", "category": "стаж", "oklad": 3500000},

    # Қийма
    {"id": "59", "fio": "Abdullayev Doniyor", "otdel": "Қийма", "category": "5 категория", "oklad": 3640000},
    {"id": "60", "fio": "Buriboev Islom", "otdel": "Қийма", "category": "3 категория", "oklad": 4900000},
    {"id": "61", "fio": "Fayzullaev Zafarjon", "otdel": "Қийма", "category": "3 категория", "oklad": 4900000},
    {"id": "62", "fio": "Malikov Azizbek", "otdel": "Қийма", "category": "стаж", "oklad": 3500000},
    {"id": "63", "fio": "Mansurov Dostonbek", "otdel": "Қийма", "category": "стаж", "oklad": 3500000},
    {"id": "64", "fio": "Orozov Faxriddin", "otdel": "Қийма", "category": "2 категория", "oklad": 6160000},
    {"id": "65", "fio": "Tadjiyev Yuldash", "otdel": "Қийма", "category": "стаж", "oklad": 3500000},

    # Котлет
    {"id": "66", "fio": "Asatullayev Sirojiddin", "otdel": "Котлет", "category": "5 категория", "oklad": 3640000},
    {"id": "67", "fio": "Shermatov Samandar", "otdel": "Котлет", "category": "3 категория", "oklad": 4900000},
    {"id": "68", "fio": "Ikromov Oybek", "otdel": "Котлет", "category": "5 категория", "oklad": 3640000},
    {"id": "69", "fio": "Habibullayev Abdullo", "otdel": "Котлет", "category": "5 категория", "oklad": 3640000},

    # Филе
    {"id": "70", "fio": "Isomitdinov Sanjarbek", "otdel": "Филе", "category": "3 категория", "oklad": 4900000},
    {"id": "71", "fio": "Xo'jakulov Bobobek", "otdel": "Филе", "category": "3 категория", "oklad": 4900000},
    {"id": "72", "fio": "Karshiboev Dilshod", "otdel": "Филе", "category": "3 категория", "oklad": 4900000},
    {"id": "73", "fio": "Jalilov Baxtiyor", "otdel": "Филе", "category": "5 категория", "oklad": 3640000},
    {"id": "74", "fio": "Toshpulatova Umida", "otdel": "Филе", "category": "3 категория", "oklad": 4900000},

    # Техперсонал
    {"id": "75", "fio": "Abduvaliev Nuriddin", "otdel": "Техперсонал", "category": "4 категория", "oklad": 4200000},
    {"id": "76", "fio": "Aminov Shahboz", "otdel": "Техперсонал", "category": "5 категория", "oklad": 3640000},
    {"id": "77", "fio": "Bo'riboev Inomjon", "otdel": "Техперсонал", "category": "5 категория", "oklad": 3640000},
    {"id": "78", "fio": "Habibullayev Doniyor", "otdel": "Техперсонал", "category": "5 категория", "oklad": 3640000},
    {"id": "79", "fio": "Isamiddinov Nuriddin", "otdel": "Техперсонал", "category": "5 категория", "oklad": 3640000},
    {"id": "80", "fio": "Xabibullayev Xayrulla", "otdel": "Техперсонал", "category": "3 категория", "oklad": 4900000},
    {"id": "81", "fio": "Jo'rayev Nodir", "otdel": "Техперсонал", "category": "нет категории", "oklad": 7000000},
    {"id": "82", "fio": "Alikulov Akmal", "otdel": "Техперсонал", "category": "4 категория", "oklad": 4200000},
    {"id": "83", "fio": "Boliyeva O'g'iloy", "otdel": "Техперсонал", "category": "5 категория", "oklad": 3640000},
    {"id": "84", "fio": "Erqo'ziev Abduazim", "otdel": "Техперсонал", "category": "5 категория", "oklad": 3640000},
    {"id": "85", "fio": "Jangirov Olimjon", "otdel": "Техперсонал", "category": "4 категория", "oklad": 4200000},
    {"id": "86", "fio": "Kuchimova Kamola", "otdel": "Техперсонал", "category": "5 категория", "oklad": 3640000},
    {"id": "87", "fio": "Qayumova Dilnoza", "otdel": "Техперсонал", "category": "5 категория", "oklad": 3640000},
    {"id": "88", "fio": "Mirzamuxammedova Nafisa", "otdel": "Техперсонал", "category": "5 категория", "oklad": 3640000},
    {"id": "89", "fio": "Samatova Yulduz", "otdel": "Техперсонал", "category": "5 категория", "oklad": 3640000},
    {"id": "90", "fio": "Narxodjayeva Kamola", "otdel": "Техперсонал", "category": "5 категория", "oklad": 3640000},
    {"id": "91", "fio": "Ruziyeva Nigora", "otdel": "Техперсонал", "category": "5 категория", "oklad": 3640000},
    {"id": "92", "fio": "Axmadjonova Zamira", "otdel": "Техперсонал", "category": "5 категория", "oklad": 3640000},
    {"id": "93", "fio": "Qodirov Baxtiyor", "otdel": "Техперсонал", "category": "нет категории", "oklad": 7500000},
    {"id": "94", "fio": "Usmonova Shaxnoza", "otdel": "Техперсонал", "category": "нет категории", "oklad": 5100000},
    {"id": "95", "fio": "Jo'raboev Muxammad", "otdel": "Техперсонал", "category": "нет категории", "oklad": 4650000},
    {"id": "96", "fio": "Karimov Abdurashid", "otdel": "Техперсонал", "category": "нет категории", "oklad": 4650000},
    {"id": "97", "fio": "Nuritdinov Jaloliddin", "otdel": "Техперсонал", "category": "нет категории", "oklad": 4650000},
    {"id": "98", "fio": "Toirov Otabek", "otdel": "Техперсонал", "category": "нет категории", "oklad": 4500000},
    {"id": "99", "fio": "Xasanxonov Azamxon", "otdel": "Техперсонал", "category": "нет категории", "oklad": 5270000},
]
# Ставки по категориям (суточная, фиксированная за день, сум)
DAILY_RATES = {
    "сихлаш 1 кат": 210000,
    "сихлаш 2 кат": 150000,
    "сихлаш 3 кат": 107000,
    "разделка 1 кат": 185000,
    "разделка 2 кат": 145000,
    "разделка 3 кат": 108000,
    "кийма 1 кат": 170000,
    "кийма 2 кат": 140000,
    "кийма 3 кат": 115000,
    "котлет 1 кат": 145000,
    "котлет 2 кат": 120000,
    "котлет 3 кат": 110000,
    "филе 1 кат": 155000,
    "филе 2 кат": 150000,
    "филе 3 кат": 100000,
    "1 категория": 214667,
    "2 категория": 205333,
    "3 категория": 163333,
    "4 категория": 140000,
    "5 категория": 121333,
    "стаж": 116667,
}

# Калибры Go'sht в кг
CALIBER_OPTIONS_KG = [5, 6, 8, 10, 12, 15, 20, 25, 30, 35, 40, 50, 60]

# Чек-лист дисциплины
CHECKLIST_CATEGORIES = [
    {"name": "Санитария и гигиена", "max_points": 20},
    {"name": "Соблюдение технологии", "max_points": 20},
    {"name": "Качество продукции", "max_points": 15},
    {"name": "Производственная дисциплина", "max_points": 15},
    {"name": "Состояние оборудования", "max_points": 10},
    {"name": "Документация и отчётность", "max_points": 10},
    {"name": "Персонал и обучение", "max_points": 10}
]

# Нормы кг по категориям
KG_NORMS = {
    "сихлаш 1 кат": 375,
    "сихлаш 2 кат": 340,
    "сихлаш 3 кат": 300,
    "разделка 1 кат": 250,
    "разделка 2 кат": 220,
    "разделка 3 кат": 200,
    "кийма 1 кат": 180,
    "кийма 2 кат": 160,
    "кийма 3 кат": 140,
    "котлет 1 кат": 120,
    "котлет 2 кат": 110,
    "котлет 3 кат": 100,
    "филе 1 кат": 150,
    "филе 2 кат": 140,
    "филе 3 кат": 130,
    "1 категория": 375,
    "2 категория": 340,
    "3 категория": 300,
    "4 категория": 250,
    "5 категория": 140,
    "стаж": 300,
}

def calculate_bonus_percent(total_points):
    if total_points >= 90:
        return 100  # 100% бонуса
    elif total_points >= 75:
        return 75
    elif total_points >= 60:
        return 50
    else:
        return 0


@app.route("/")
def index():
    """Главная страница — редирект на табель"""
    return redirect("/tabel")


@app.route("/input", methods=["GET", "POST"])
def input_prod():
    success = None
    error = None

    if request.method == "POST":
        worker_id = request.form.get("worker_id")
        fio = request.form.get("fio")
        product = request.form.get("product")
        quantity = request.form.get("quantity")
        caliber_kg = request.form.get("caliber_kg")
        date = request.form.get("date")

        if not fio or not product or not quantity or not caliber_kg or not date:
            error = "Заполните все поля!"
        else:
            try:
                quantity_pieces = int(quantity)
                selected_caliber_kg = float(caliber_kg)
                quantity_kg = quantity_pieces * selected_caliber_kg

                # Находим работника по ID или ФИО
                worker = next((w for w in WORKERS_TABLE if w["id"] == worker_id or w["fio"] == fio), None)

                if not worker:
                    error = "Ошибка: работник не найден!"
                else:
                    category = worker.get("category", "")
                    daily_rate = DAILY_RATES.get(category, 0)
                    daily_salary = daily_rate  # фиксированная ставка за день

                    # Расчет нормы и % выполнения
                    norm_kg = KG_NORMS.get(category, 0)
                    percent_complete = (quantity_kg / norm_kg * 100) if norm_kg > 0 else 0

                    # Расчет коэффициента з/п и бонуса по норме
                    if percent_complete >= 90:
                        salary_coeff = 1.0
                        if percent_complete >= 100:
                            bonus_percent = 20
                        elif percent_complete >= 80:
                            bonus_percent = 10
                        else:
                            bonus_percent = 0
                    elif percent_complete >= 80:
                        salary_coeff = 0.9
                        bonus_percent = 0
                    elif percent_complete >= 70:
                        salary_coeff = 0.7
                        bonus_percent = 0
                    else:
                        salary_coeff = 0.5
                        bonus_percent = 0

                    daily_salary = int(daily_rate * salary_coeff)

                    full_salary = daily_rate  # полная ставка без коэффициента
                    reduced_amount = full_salary - daily_salary  # сколько урезано

                    # Чек-лист дисциплины
                    discipline_total = int(request.form.get('discipline_total', 0))
                    
                    # Отдел работника
                    otdel = worker.get("otdel", "")
                    record = {
                        "worker_id": worker["id"],
                        "fio": fio,
                        "product": product,
                        "quantity_pieces": quantity_pieces,
                        "caliber_kg": selected_caliber_kg,
                        "quantity_kg": round(quantity_kg, 1),
                        "category": category,
                        "daily_rate": daily_rate,
                        "salary_coeff": salary_coeff,
                        "daily_salary": daily_salary,
                        "full_salary": full_salary,
                        "reduced_amount": reduced_amount,
                        "bonus_percent": bonus_percent,
                        "otdel": otdel,
                        "date": date,
                        "source": "manual",
                        "discipline_total": discipline_total,
                        "norm_kg": norm_kg,
                        "percent_complete": round(percent_complete, 1)
                    }
                    add_record(record)
                    success = f"Добавлено: {fio} — {quantity_pieces} шт × {selected_caliber_kg} кг = {quantity_kg} кг | З/П: {daily_salary:,} сум"
            except ValueError:
                error = "Ошибка при расчете количества!"

    return render_template("input.html", 
                          groups=GROUPS_LIST,
                          products_by_group=PRODUCT_GROUPS,
                          workers=WORKERS_TABLE,
                          caliber_options=CALIBER_OPTIONS_KG,
                          daily_rates=DAILY_RATES,
                          success=success, 
                          error=error,
                          kg_norms=KG_NORMS)




@app.template_filter('format_sum')
def format_sum(value):
    """Форматирует число с пробелами как разделитель тысяч"""
    return "{:,}".format(int(value)).replace(",", " ")


@app.route("/delete/<int:index>")
def delete_record(index):
    """Удаляет запись по индексу"""
    import sqlite3
    from database import DB_FILE
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("DELETE FROM records WHERE id = ?", (index,))
    conn.commit()
    conn.close()
    return redirect("/tabel")


@app.route('/update_date/<int:record_id>', methods=['POST'])
def update_date(record_id):
    """Обновляет дату записи"""
    import sqlite3
    from database import DB_FILE
    from flask import request, jsonify
    
    data = request.get_json()
    new_date = data.get('date')
    
    if not new_date:
        return jsonify({'error': 'Дата не указана'}), 400
    
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("UPDATE records SET date = ? WHERE id = ?", (new_date, record_id))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True})


@app.route("/tabel")
def tabel():
    otdel_filter = request.args.get("otdel", "Все")
    fio_filter = request.args.get("fio", "Все").strip()
    date_filter = request.args.get("date", "")
    
    print(f"DEBUG: otdel_filter = '{otdel_filter}', fio_filter = '{fio_filter}', date_filter = '{date_filter}'")
    
    # Получаем записи из базы данных
    if date_filter:
        all_records = get_records_by_date(date_filter)
    else:
        all_records = get_all_records()
    
    filtered_prod = all_records.copy()
    
    # Фильтрация по отделу
    if otdel_filter != "Все":
        filtered_prod = [r for r in filtered_prod if r.get("otdel") == otdel_filter]
        print(f"DEBUG: После фильтрации по отделу '{otdel_filter}': {len(filtered_prod)} записей")
    
    # Фильтрация по ФИО или ID
    if fio_filter != "Все" and fio_filter != "":
        # Если ввели цифры - ищем по ID и подставляем ФИО
        if fio_filter.isdigit():
            worker_by_id = next((w for w in WORKERS_TABLE if str(w["id"]) == fio_filter), None)
            if worker_by_id:
                fio_filter = worker_by_id["fio"]
                print(f"DEBUG: Найден работник по ID {fio_filter}: {fio_filter}")
        
        # Фильтруем по ФИО
        filtered_prod = [r for r in filtered_prod if r.get("fio") == fio_filter]
        print(f"DEBUG: После фильтрации по ФИО '{fio_filter}': {len(filtered_prod)} записей")
    
    # Проверка несоответствия работника отделу
    warning_message = None
    if fio_filter != "Все" and fio_filter != "" and otdel_filter != "Все":
        selected_worker = next((w for w in WORKERS_TABLE if w["fio"] == fio_filter), None)
        if selected_worker and selected_worker["otdel"] != otdel_filter:
            warning_message = f'"{fio_filter}" — работник отдела "{selected_worker["otdel"]}", а не "{otdel_filter}"'
    
    # Рассчитываем итоги за день/фильтр
    total_records = len(filtered_prod)
    total_kg = sum(r.get("quantity_kg", 0) for r in filtered_prod)
    total_salary = sum(r.get("daily_salary", 0) for r in filtered_prod)
    
    # Итого за месяц (все даты)
    all_records_for_month = get_all_records()
    monthly_salary = sum(r.get("daily_salary", 0) for r in all_records_for_month)
    
    total_salary_formatted = "{:,}".format(int(total_salary)).replace(",", " ")
    monthly_salary_formatted = "{:,}".format(int(monthly_salary)).replace(",", " ")
    
    has_records = len(filtered_prod) > 0
    print(f"DEBUG: has_records = {has_records}")
    
    return render_template("tabel.html", 
                          workers=WORKERS_TABLE,
                          daily_prod=filtered_prod,
                          category_rates=DAILY_RATES,
                          total_records=total_records,
                          total_kg=total_kg,
                          total_salary=total_salary,
                          monthly_salary=monthly_salary,
                          total_salary_formatted=total_salary_formatted,
                          monthly_salary_formatted=monthly_salary_formatted,
                          otdel_filter=otdel_filter,
                          fio_filter=fio_filter,
                          has_records=has_records,
                          warning_message=warning_message)


@app.route("/export")
def export_excel():
    date_filter = request.args.get("date")

    if date_filter:
        filtered = get_records_by_date(date_filter)
        filename = f"tabell_{date_filter}.xlsx"
    else:
        filtered = get_all_records()
        filename = "tabell_vse.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Табель производства"

    # Заголовок
    ws['A1'] = "Табель производства"
    ws.merge_cells('A1:N1')
    ws['A1'].font = Font(bold=True, size=14)

    # Заголовок даты
    ws['A2'] = f"Дата: {date_filter or 'Все даты'}"
    ws.merge_cells('A2:N2')
    ws['A2'].font = Font(size=12)

    # Столбцы
    headers = ["ID", "ФИО", "Товар", "Категория", "Кол-во (шт)", "Калибр (кг)", "Общий вес (кг)", "З/п полная (сум)", "З/п за смену (сум)", "Урезано (сум)", "Коэф. з/п", "Баллы всего", "Бонус %", "Норма (кг)", "Выполнено (%)", "Источник", "Дата"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Данные
    row = 4
    for record in filtered:
        ws.append([
            record.get("worker_id", ""),
            record.get("fio", ""),
            record.get("product", ""),
            record.get("category", ""),
            record.get("quantity_pieces", 0),
            record.get("caliber_kg", 0),
            record.get("quantity_kg", 0),
            record.get("full_salary", 0),
            record.get("daily_salary", 0),
            record.get("reduced_amount", 0),
            record.get("salary_coeff", 1.0),
            record.get("total_points", 0),
            record.get("bonus_percent", 0),
            record.get("norm_kg", 0),
            record.get("percent_complete", 0),
            record.get("source", "Вручную"),
            record.get("date", "")
        ])
        row += 1

    # Итоги
    total_kg = sum(r.get("quantity_kg", 0) for r in filtered)
    total_full_salary = sum(r.get("full_salary", 0) for r in filtered)
    total_salary_sum = sum(r.get("daily_salary", 0) for r in filtered)
    total_reduced = sum(r.get("reduced_amount", 0) for r in filtered)

    ws.append([])
    # Строка итогов должна иметь столько же столбцов, сколько headers
    totals_row_values = [
        "Итого:", "", "", "", "", "",
        total_kg,
        total_full_salary,
        total_salary_sum,
        total_reduced,
        "", "", "", "", "", "", ""
    ]
    ws.append(totals_row_values)
    totals_row_index = row + 1
    ws.cell(row=totals_row_index, column=7).font = Font(bold=True)
    ws.cell(row=totals_row_index, column=7).number_format = '#,##0.0'
    ws.cell(row=totals_row_index, column=8).font = Font(bold=True)
    ws.cell(row=totals_row_index, column=8).number_format = '#,##0'
    ws.cell(row=totals_row_index, column=9).font = Font(bold=True)
    ws.cell(row=totals_row_index, column=9).number_format = '#,##0'
    ws.cell(row=totals_row_index, column=10).font = Font(bold=True)
    ws.cell(row=totals_row_index, column=10).number_format = '#,##0'

    # Автоширина столбцов
    for col in range(1, len(headers)+1):
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Отладка
    print("Экспорт: дата фильтра =", date_filter)
    print("Количество записей для экспорта =", len(filtered))

    # Сохраняем в память
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


@app.route("/monthly_report", methods=["GET", "POST"])
def monthly_report():
    if request.method == "POST":
        month = request.form.get("month")  # формат YYYY-MM
        if not month:
            month = datetime.now().strftime("%Y-%m")

        # Фильтр записей за месяц
        year, mon = map(int, month.split("-"))
        all_records = get_all_records()
        monthly_records = [
            r for r in all_records
            if r["date"].startswith(f"{year}-{mon:02d}")
        ]

        # Группировка: сначала по отделу, затем по ФИО
        by_otdel = defaultdict(lambda: defaultdict(lambda: {
            "total_kg": 0,
            "total_salary": 0,
            "count": 0,
            "total_complete": 0,
            "total_points": 0
        }))

        for r in monthly_records:
            otdel = r.get("otdel", "Неизвестно")
            fio = r.get("fio", "Неизвестно")
            data = by_otdel[otdel][fio]
            data["total_kg"] += r.get("quantity_kg", 0)
            data["total_salary"] += r.get("daily_salary", 0)
            data["count"] += 1
            data["total_complete"] += r.get("percent_complete", 0)
            data["total_points"] += r.get("total_points", 0)

        # Формируем отчёт по отделам и работникам
        report = []
        grand_total_kg = 0
        grand_total_salary = 0

        for otdel, workers in by_otdel.items():
            otdel_total_kg = 0
            otdel_total_salary = 0
            otdel_workers = []

            for fio, data in workers.items():
                avg_complete = data["total_complete"] / data["count"] if data["count"] > 0 else 0
                avg_bonus = data["total_points"] / data["count"] if data["count"] > 0 else 0

                otdel_workers.append({
                    "fio": fio,
                    "total_kg": round(data["total_kg"], 1),
                    "total_salary": int(data["total_salary"]),
                    "avg_complete": round(avg_complete, 1),
                    "avg_bonus": round(avg_bonus, 1)
                })

                otdel_total_kg += data["total_kg"]
                otdel_total_salary += data["total_salary"]

            report.append({
                "otdel": otdel,
                "workers": otdel_workers,
                "total_kg": round(otdel_total_kg, 1),
                "total_salary": int(otdel_total_salary)
            })

            grand_total_kg += otdel_total_kg
            grand_total_salary += otdel_total_salary

        return render_template(
            "monthly_report.html",
            month=month,
            report=report,
            grand_total_kg=round(grand_total_kg, 1),
            grand_total_salary=grand_total_salary
        )

    # GET — форма выбора месяца (пока без отчёта)
    current_month = datetime.now().strftime("%Y-%m")
    return render_template(
        "monthly_report.html",
        month=current_month,
        report=None,
        grand_total_kg=0,
        grand_total_salary=0
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    debug_mode = os.environ.get("DEBUG", "False") == "True"
    app.run(host="0.0.0.0", port=port, debug=debug_mode)
